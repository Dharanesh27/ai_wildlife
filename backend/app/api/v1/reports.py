import io
from uuid import UUID
from fastapi import APIRouter, Depends, status, HTTPException
from fastapi.responses import StreamingResponse
from app.api.dependencies import (
    get_current_user,
    get_survey_site_repository,
    get_device_repository,
    get_observation_repository,
    get_ecosystem_health_log_repository,
    get_recommendation_repository,
)
from app.domain.models.user import User
from app.repositories.survey_repo import (
    SurveySiteRepository,
    DeviceRepository,
    ObservationRepository,
    EcosystemHealthLogRepository,
    RecommendationRepository,
)

router = APIRouter()

# Authentication Protection
allow_all_authenticated = Depends(get_current_user)

@router.get("/pdf/{site_id}")
async def export_pdf_report(
    site_id: UUID,
    site_repo: SurveySiteRepository = Depends(get_survey_site_repository),
    device_repo: DeviceRepository = Depends(get_device_repository),
    observation_repo: ObservationRepository = Depends(get_observation_repository),
    health_repo: EcosystemHealthLogRepository = Depends(get_ecosystem_health_log_repository),
    rec_repo: RecommendationRepository = Depends(get_recommendation_repository),
    current_user: User = allow_all_authenticated,
):
    """Generates and downloads a compiled PDF reserve intelligence report."""
    # 1. Fetch data
    site = await site_repo.get(site_id)
    if not site:
        raise HTTPException(status_code=404, detail="Survey site not found")
        
    devices = await device_repo.get_by_site(site_id)
    observations = await observation_repo.get_by_site(site_id, limit=20)
    health_logs = await health_repo.get_by_site(site_id, limit=5)
    recommendations = await rec_repo.get_by_site(site_id)

    # 2. Build PDF Document using ReportLab
    from reportlab.lib.pagesizes import letter
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        rightMargin=40, leftMargin=40, topMargin=40, bottomMargin=40
    )
    
    story = []
    styles = getSampleStyleSheet()
    
    # Custom Modern Styles
    title_style = ParagraphStyle(
        'ReportTitle',
        parent=styles['Heading1'],
        fontSize=22,
        leading=26,
        textColor=colors.HexColor('#0f172a'), # slate-900
        spaceAfter=15
    )
    section_style = ParagraphStyle(
        'SectionHeader',
        parent=styles['Heading2'],
        fontSize=14,
        leading=18,
        textColor=colors.HexColor('#10b981'), # emerald-500
        spaceBefore=15,
        spaceAfter=10
    )
    body_style = ParagraphStyle(
        'ReportBody',
        parent=styles['Normal'],
        fontSize=10,
        leading=14,
        textColor=colors.HexColor('#334155') # slate-700
    )
    bold_body_style = ParagraphStyle(
        'ReportBodyBold',
        parent=body_style,
        fontName='Helvetica-Bold'
    )
    table_header_style = ParagraphStyle(
        'TableHeader',
        parent=body_style,
        textColor=colors.white,
        fontName='Helvetica-Bold'
    )

    # Title & Metadata Banner
    story.append(Paragraph("AI Wildlife Population Intelligence Report", title_style))
    story.append(Paragraph(f"Generated on: {health_logs[0].logged_at.strftime('%Y-%m-%d %H:%M:%S UTC') if health_logs else 'Live Data'}", body_style))
    story.append(Spacer(1, 15))

    # Reserve Overview
    story.append(Paragraph("1. Reserve Scope Overview", section_style))
    site_data = [
        [Paragraph("Reserve Name:", bold_body_style), Paragraph(site.name, body_style)],
        [Paragraph("Location Region:", bold_body_style), Paragraph(site.location_name, body_style)],
        [Paragraph("Habitat Profile:", bold_body_style), Paragraph(site.habitat_type, body_style)],
        [Paragraph("Coordinates HUD:", bold_body_style), Paragraph(f"{site.latitude:.5f} N, {site.longitude:.5f} E", body_style)],
        [Paragraph("Protection Status:", bold_body_style), Paragraph("Designated Protected Reserve" if site.is_protected_area else "Public/Buffer Buffer Zone", body_style)],
    ]
    site_table = Table(site_data, colWidths=[120, 400])
    site_table.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 4),
        ('LINEBELOW', (0,0), (-1,-1), 0.5, colors.HexColor('#e2e8f0')),
    ]))
    story.append(site_table)
    story.append(Spacer(1, 20))

    # Ecosystem Health Scoring
    story.append(Paragraph("2. Ecosystem Health Index", section_style))
    if health_logs:
        latest = health_logs[0]
        health_data = [
            [Paragraph("Indicator Metric", table_header_style), Paragraph("Weight", table_header_style), Paragraph("Score", table_header_style)],
            [Paragraph("Species Diversity", body_style), Paragraph("30%", body_style), Paragraph(f"{latest.biodiversity_score:.2f} / 10", body_style)],
            [Paragraph("Population Stability", body_style), Paragraph("25%", body_style), Paragraph(f"{latest.population_stability_score:.2f} / 10", body_style)],
            [Paragraph("Habitat Quality", body_style), Paragraph("20%", body_style), Paragraph(f"{latest.habitat_quality_score:.2f} / 10", body_style)],
            [Paragraph("Endangered Species Status", body_style), Paragraph("15%", body_style), Paragraph(f"{latest.endangered_species_status_score:.2f} / 10", body_style)],
            [Paragraph("Environmental Conditions", body_style), Paragraph("10%", body_style), Paragraph(f"{latest.environmental_conditions_score:.2f} / 10", body_style)],
            [Paragraph("Overall Ecosystem Health Index", bold_body_style), Paragraph("100%", bold_body_style), Paragraph(f"{latest.overall_health_score:.2f} / 10", bold_body_style)],
        ]
        health_table = Table(health_data, colWidths=[200, 100, 220])
        health_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1e293b')),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('BOTTOMPADDING', (0,0), (-1,-1), 5),
            ('TOPPADDING', (0,0), (-1,-1), 5),
            ('LINEBELOW', (0,0), (-1,-1), 0.5, colors.HexColor('#cbd5e1')),
            ('BACKGROUND', (0,-1), (-1,-1), colors.HexColor('#f1f5f9')),
        ]))
        story.append(health_table)
    else:
        story.append(Paragraph("Ecosystem health metrics not calibrated yet. Recalculate health on the dashboard to populate.", body_style))
    story.append(Spacer(1, 20))

    # Hardware Telemetry Stats
    story.append(Paragraph("3. Hardware Telemetry Coverage", section_style))
    camera_count = len([d for d in devices if d.device_type.value == "Camera Trap"])
    audio_count = len([d for d in devices if d.device_type.value == "Audio Sensor"])
    story.append(Paragraph(f"Active monitoring hardware deployed: <b>{len(devices)} stations</b> ({camera_count} Camera Traps, {audio_count} Acoustic Sensors).", body_style))
    story.append(Spacer(1, 15))

    # Recent Observations
    story.append(Paragraph("4. Recent Telemetry Sighting Log", section_style))
    if observations:
        obs_data = [
            [Paragraph("Timestamp (UTC)", table_header_style), Paragraph("Device", table_header_style), Paragraph("Detected Species", table_header_style), Paragraph("Confidence", table_header_style), Paragraph("Threat", table_header_style)]
        ]
        
        # Map device ids to names for readability
        dev_map = {d.id: d.name for d in devices}
        
        for obs in observations[:12]:
            dev_name = dev_map.get(obs.device_id, "Unknown Node")
            threat_color = "#f43f5e" if obs.threat_level.value in ["High", "Critical"] else "#334155"
            threat_paragraph = Paragraph(f"<font color='{threat_color}'><b>{obs.threat_level.value}</b></font>", body_style)
            
            obs_data.append([
                Paragraph(obs.timestamp.strftime('%Y-%m-%d %H:%M'), body_style),
                Paragraph(dev_name, body_style),
                Paragraph(obs.detected_species, body_style),
                Paragraph(f"{obs.confidence:.1f}%", body_style),
                threat_paragraph
            ])
            
        obs_table = Table(obs_data, colWidths=[110, 110, 160, 70, 70])
        obs_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1e293b')),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('BOTTOMPADDING', (0,0), (-1,-1), 4),
            ('TOPPADDING', (0,0), (-1,-1), 4),
            ('LINEBELOW', (0,0), (-1,-1), 0.5, colors.HexColor('#cbd5e1')),
        ]))
        story.append(obs_table)
    else:
        story.append(Paragraph("No sightings recorded in telemetry history log.", body_style))
    story.append(Spacer(1, 20))

    # Active Recommendations / Patrol Directives
    story.append(Paragraph("5. Active Patrol & Tactical Directives", section_style))
    open_recs = [r for r in recommendations if r.status.value == "Open"]
    if open_recs:
        for rec in open_recs[:5]:
            rec_priority_color = "#f43f5e" if rec.priority.value == "Critical" else "#f59e0b" if rec.priority.value == "Medium" else "#10b981"
            rec_head = f"<b>{rec.title}</b> (<font color='{rec_priority_color}'><b>{rec.priority.value} Priority</b></font>)"
            story.append(Paragraph(rec_head, body_style))
            story.append(Paragraph(f"Details: {rec.description}", body_style))
            story.append(Spacer(1, 8))
    else:
        story.append(Paragraph("No open patrol alerts. Current reserve coverage sector is secure.", body_style))

    # Build PDF
    doc.build(story)
    buffer.seek(0)
    
    clean_filename = f"Wildlife_Report_{site.name.replace(' ', '_')}.pdf"
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={clean_filename}"}
    )


@router.get("/excel/{site_id}")
async def export_excel_report(
    site_id: UUID,
    site_repo: SurveySiteRepository = Depends(get_survey_site_repository),
    device_repo: DeviceRepository = Depends(get_device_repository),
    observation_repo: ObservationRepository = Depends(get_observation_repository),
    current_user: User = allow_all_authenticated,
):
    """Generates and downloads an Excel workbook containing device status grids and observation records."""
    site = await site_repo.get(site_id)
    if not site:
        raise HTTPException(status_code=404, detail="Survey site not found")
        
    devices = await device_repo.get_by_site(site_id)
    observations = await observation_repo.get_by_site(site_id, limit=5000)

    # Build Excel Workbook using openpyxl
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    wb = openpyxl.Workbook()
    
    # ----------------------------------------------------
    # Sheet 1: Observations Log
    # ----------------------------------------------------
    ws1 = wb.active
    ws1.title = "Observations Log"
    
    # Header styling
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid") # slate-800
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    center_align = Alignment(horizontal="center", vertical="center")
    
    headers1 = [
        "Sighting ID", "Timestamp (UTC)", "Device Station", 
        "Observation Type", "Detected Species", "Taxonomic Class", 
        "Sighting Count", "Observed Behavior", "Confidence %", 
        "Threat Level", "Threat Details", "File Resource URL"
    ]
    
    ws1.append(headers1)
    for col_num, header in enumerate(headers1, 1):
        cell = ws1.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    # Device mapping for easy lookup
    dev_map = {d.id: d.name for d in devices}

    # Populate Data
    for obs in observations:
        dev_name = dev_map.get(obs.device_id, "Unknown Node")
        ws1.append([
            str(obs.id),
            obs.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
            dev_name,
            obs.observation_type.value,
            obs.detected_species,
            obs.taxonomic_class,
            obs.count,
            obs.behavior or "",
            obs.confidence,
            obs.threat_level.value,
            obs.threat_details or "",
            obs.file_url or ""
        ])

    # Adjust column widths
    for col in ws1.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws1.column_dimensions[col_letter].width = max(max_len + 3, 11)

    # ----------------------------------------------------
    # Sheet 2: Hardware Devices
    # ----------------------------------------------------
    ws2 = wb.create_sheet(title="Hardware Devices")
    
    headers2 = [
        "Device ID", "Station Name", "Device Type", 
        "Current Status", "Battery Level %", "Latitude (WGS84)", "Longitude (WGS84)"
    ]
    
    ws2.append(headers2)
    for col_num, header in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    for dev in devices:
        ws2.append([
            str(dev.id),
            dev.name,
            dev.device_type.value,
            dev.status.value,
            dev.battery_level,
            dev.latitude,
            dev.longitude
        ])

    for col in ws2.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws2.column_dimensions[col_letter].width = max(max_len + 3, 11)

    # Save to BytesIO
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    clean_filename = f"Wildlife_Telemetry_{site.name.replace(' ', '_')}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={clean_filename}"}
    )
